"""소스 자료 추출 → /tmp/ubi-lang-raw/ 에 저장.

엑셀(openpyxl), PDF(pdfplumber), MVP(파일 읽기)에서 도메인 용어 후보가 될
텍스트 데이터를 뽑는다. raw 데이터는 임시 폴더에 두고 git에 포함하지 않음.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber

RAW_DIR = Path("/tmp/ubi-lang-raw")


def extract_excel(path: Path) -> dict[str, Any]:
    """엑셀 1개 → 시트별 메타데이터.

    각 시트에서 추출:
    - name: 시트 이름
    - headers: 1행 값 (헤더 추정)
    - text_cells: 모든 문자열 셀 (중복 제거, 순서 유지)
    - merged_cells: 결합셀 범위 (제목/소계 구조)
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        headers = []
        if ws.max_row >= 1:
            for cell in ws[1]:
                v = cell.value
                if v is not None:
                    headers.append(str(v).strip())
        seen: set[str] = set()
        text_cells: list[str] = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    s = v.strip()
                    if s and s not in seen:
                        seen.add(s)
                        text_cells.append(s)
        merged = [str(r) for r in ws.merged_cells.ranges]
        sheets.append({
            "name": ws.title,
            "headers": headers,
            "text_cells": text_cells,
            "merged_cells": merged,
        })
    wb.close()
    return {"path": str(path), "sheets": sheets}


def extract_pdf(path: Path) -> dict[str, Any]:
    """PDF → 페이지별 텍스트.

    각 페이지에서 추출:
    - page_number: 1-based
    - text: 페이지 전체 텍스트 (pdfplumber.extract_text)
    - tables: 페이지의 표 (있다면 list[list[list[str]]])
    """
    pages = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            tables = page.extract_tables() or []
            pages.append({
                "page_number": i,
                "text": text,
                "tables": tables,
            })
    return {"path": str(path), "pages": pages}


MVP_ROOT = Path("/Users/apple/Desktop/git/see-architect-mvp")


def extract_mvp() -> dict[str, Any]:
    """MVP 저장소의 도메인 관련 파일 발췌.

    포함:
    - CLAUDE.md 전체
    - src/store/types.ts 전체
    - src/domain/*.ts 전체 (deriveContract, weeklyMath, dashboardStats 등)
    """
    claude_md = (MVP_ROOT / "CLAUDE.md").read_text(encoding="utf-8")
    types_ts = (MVP_ROOT / "src/store/types.ts").read_text(encoding="utf-8")
    domain_dir = MVP_ROOT / "src/domain"
    domain_files: dict[str, str] = {}
    for p in sorted(domain_dir.glob("*.ts")):
        if p.name.endswith(".test.ts"):
            continue
        domain_files[p.name] = p.read_text(encoding="utf-8")
    return {
        "claude_md": claude_md,
        "types_ts": types_ts,
        "domain_files": domain_files,
    }


SRC_DIR = Path("/Users/apple/Desktop/project/see-architect/files")

EXCEL_FILES = {
    "신사동": "2413 신사동 626-77 근생 .xlsx",
    "지급": "지급내역서_샘플.xlsx",
    "직접간접": "현장_직접비_및_간접비_내역.xlsx",
    "경비": "현장소장_월_현장경비_샘플.xlsx",
}

PDF_FILE = "01. 제 4회 하이콘 뉴캠퍼스 C동 리모델링공사 기성청구서(합본).pdf"


def main() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    (RAW_DIR / "excel").mkdir(exist_ok=True)
    (RAW_DIR / "pdf").mkdir(exist_ok=True)
    (RAW_DIR / "mvp").mkdir(exist_ok=True)

    for alias, filename in EXCEL_FILES.items():
        print(f"[excel] {alias} ← {filename}")
        result = extract_excel(SRC_DIR / filename)
        out = RAW_DIR / "excel" / f"{alias}.json"
        out.write_text(json.dumps(result, ensure_ascii=False, indent=2),
                       encoding="utf-8")
        print(f"  → {out} ({len(result['sheets'])} sheets)")

    print(f"[pdf] 기성 ← {PDF_FILE}")
    pdf_result = extract_pdf(SRC_DIR / PDF_FILE)
    pdf_out = RAW_DIR / "pdf" / "기성.json"
    pdf_out.write_text(json.dumps(pdf_result, ensure_ascii=False, indent=2),
                       encoding="utf-8")
    print(f"  → {pdf_out} ({len(pdf_result['pages'])} pages)")

    print("[mvp] CLAUDE.md, types.ts, domain/*.ts")
    mvp = extract_mvp()
    mvp_md_lines = [
        "# MVP 발췌\n",
        "## CLAUDE.md\n", "```markdown", mvp["claude_md"], "```\n",
        "## src/store/types.ts\n", "```typescript", mvp["types_ts"], "```\n",
        "## src/domain/\n",
    ]
    for name, content in mvp["domain_files"].items():
        mvp_md_lines += [f"### {name}\n", "```typescript", content, "```\n"]
    (RAW_DIR / "mvp" / "extracted.md").write_text(
        "\n".join(mvp_md_lines), encoding="utf-8"
    )
    print(f"  → {RAW_DIR / 'mvp' / 'extracted.md'}")
    print("\n추출 완료.")


if __name__ == "__main__":
    main()
